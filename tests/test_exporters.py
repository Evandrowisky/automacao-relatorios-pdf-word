"""Testes dos exportadores TXT, CSV e Excel."""

import csv
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


SRC_DIR = Path(__file__).resolve().parents[1] / "src"
sys.path.append(str(SRC_DIR))

from csv_exporter import export_rows_to_csv  # noqa: E402
from excel_exporter import export_rows_to_excel  # noqa: E402
from text_exporter import build_text_content, export_text  # noqa: E402


class ExportersTest(unittest.TestCase):
    """Valida a criação dos principais formatos de saída."""

    def test_export_text_creates_txt_with_page_separators(self) -> None:
        extraction = {
            "arquivo": "documento.pdf",
            "paginas": [
                {"numero": 1, "texto": "Texto da primeira página."},
                {"numero": 2, "texto": "Texto da segunda página."},
            ],
            "texto_completo": "Texto da primeira página.\nTexto da segunda página.",
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = export_text(extraction, Path(temp_dir))
            content = output_path.read_text(encoding="utf-8")

        self.assertIn("ARQUIVO DE ORIGEM: documento.pdf", content)
        self.assertIn("================ PÁGINA 1 ================", content)
        self.assertIn("================ PÁGINA 2 ================", content)

    def test_build_text_content_returns_utf8_ready_content(self) -> None:
        extraction = {
            "arquivo": "origem.pdf",
            "paginas": [{"numero": 1, "texto": "Conteúdo com acentuação."}],
            "texto_completo": "Conteúdo com acentuação.",
        }

        content = build_text_content(extraction)

        self.assertIn("Conteúdo com acentuação.", content)

    def test_export_rows_to_csv_uses_semicolon(self) -> None:
        rows = [{"arquivo_origem": "a.pdf", "titulo": "Teste"}]

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "dados.csv"
            export_rows_to_csv(rows, output_path, ["titulo"])

            with output_path.open(encoding="utf-8-sig", newline="") as csv_file:
                reader = csv.DictReader(csv_file, delimiter=";")
                loaded_rows = list(reader)

        self.assertEqual(loaded_rows[0]["arquivo_origem"], "a.pdf")
        self.assertEqual(loaded_rows[0]["titulo"], "Teste")

    def test_export_rows_to_excel_creates_formatted_workbook(self) -> None:
        rows = [{"arquivo_origem": "a.pdf", "titulo": "Teste"}]
        summary = {"PDFs processados": 1, "Arquivos com sucesso": 1}

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "dados.xlsx"
            export_rows_to_excel(rows, output_path, ["titulo"], summary)
            workbook = load_workbook(output_path)

        self.assertIn("Dados extraídos", workbook.sheetnames)
        self.assertIn("Resumo da execução", workbook.sheetnames)
        self.assertEqual(workbook["Dados extraídos"]["A1"].value, "arquivo_origem")
        self.assertTrue(workbook["Dados extraídos"]["A1"].font.bold)


if __name__ == "__main__":
    unittest.main()
