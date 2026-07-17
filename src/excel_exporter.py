"""Exportação de dados estruturados para Excel."""

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


LOGGER = logging.getLogger(__name__)


def export_rows_to_excel(
    rows: list[dict[str, str]],
    output_path: Path,
    field_names: list[str],
    execution_summary: dict[str, Any],
) -> Path:
    """Salva os dados extraídos em uma planilha XLSX formatada."""

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Dados extraídos"

    columns = ["arquivo_origem", *field_names]
    write_data_sheet(data_sheet, columns, rows)
    write_summary_sheet(workbook, execution_summary)

    workbook.save(output_path)
    LOGGER.info("Arquivo Excel gerado: %s", output_path)

    return output_path


def write_data_sheet(sheet, columns: list[str], rows: list[dict[str, str]]) -> None:
    """Preenche e formata a planilha principal."""

    sheet.append(columns)

    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    format_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    format_cells(sheet)
    adjust_column_widths(sheet)


def write_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    """Cria a planilha de resumo da execução."""

    sheet = workbook.create_sheet("Resumo da execução")
    sheet.append(["Métrica", "Valor"])

    for key, value in summary.items():
        sheet.append([key, value])

    format_header(sheet)
    adjust_column_widths(sheet)


def format_header(sheet) -> None:
    """Aplica negrito no cabeçalho da planilha."""

    for cell in sheet[1]:
        cell.font = Font(bold=True)


def format_cells(sheet) -> None:
    """Habilita quebra de linha e alinhamento vertical."""

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def adjust_column_widths(sheet) -> None:
    """Ajusta a largura das colunas com limite para textos longos."""

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
